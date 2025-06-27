import os, sys, json, platform, threading, queue, time
from datetime import datetime, timedelta
from openpyxl import load_workbook, Workbook
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import ttkbootstrap as tb  # pip install ttkbootstrap
from ttkbootstrap.constants import *

# ── RESOURCE LOADER FOR PyInstaller ────────────────────
def resource_path(rel):
    try:
        base = sys._MEIPASS
    except AttributeError:
        base = os.path.abspath('.')
    return os.path.join(base, rel)

# ── CONFIG ─────────────────────────────────────────────
with open(resource_path('config.json')) as f:
    cfg = json.load(f)
NUM_STATIONS = int(cfg.get('num_stations', 8))
DATA_FOLDER = 'data'
TEMPLATE_PATH = resource_path('templates/excel_template.xlsx')
CHECKPOINTS = [20, 40]  # seconds for testing
LOG_PATH = os.path.join(DATA_FOLDER, datetime.now().strftime('%Y%m%d') + '.xlsx')

os.makedirs(DATA_FOLDER, exist_ok=True)
if not os.path.exists(LOG_PATH):
    from shutil import copyfile
    copyfile(TEMPLATE_PATH, LOG_PATH)

# ── UTILS ──────────────────────────────────────────────
def center(win, w=None, h=None):
    win.update_idletasks()
    width = w or win.winfo_width()
    height = h or win.winfo_height()
    x = (win.winfo_screenwidth() - width) // 2
    y = (win.winfo_screenheight() - height) // 2
    win.geometry(f"{width}x{height}+{x}+{y}")

def beep():
    if platform.system() == 'Windows':
        import winsound; winsound.Beep(1000, 300)
    else:
        print('\a', end='', flush=True)

# ── DATA MODEL ─────────────────────────────────────────
class TimerData:
    def __init__(self, serial, station, tech):
        self.serial = serial
        self.station = station
        self.tech = tech
        self.start = datetime.now()
        self.done_checks = 0
        self.cancelled = False
        self.error = False

    def next_delta(self):
        if self.done_checks >= len(CHECKPOINTS):
            return None
        target = self.start + timedelta(seconds=CHECKPOINTS[self.done_checks])
        return (target - datetime.now()).total_seconds()

    def status(self):
        d = self.next_delta()
        if d is None:
            return '-', ''
        if d < 0:
            s = int(-d)
            return 'OVERDUE', f'{s//60}:{s%60:02d}'
        m, s = divmod(int(d), 60)
        return f'{m}:{s:02d}', ''

# ── GUI SETUP ───────────────────────────────────────────
root = tb.Window(themename='litera')
root.title('Device Logger')
root.attributes('-fullscreen', True)

serial_var = tk.StringVar()
tech_var = tk.StringVar()
timers = {}
prompt_q = queue.Queue()
prompt_lock = threading.Semaphore(1)
activity_log = []

# ── TREEVIEW ───────────────────────────────────────────
cols = ('Station','Serial','Tech','Progress','Next','Error','Overdue')
tree = ttk.Treeview(root, columns=cols, show='headings')
for c in cols:
    tree.heading(c, text=c)
    tree.column(c, anchor='center')
tree.pack(fill='both', expand=True)

def refresh_tree():
    sel = tree.selection()
    sel_st = tree.item(sel[0])['values'][0] if sel else None
    tree.delete(*tree.get_children())
    for st, td in sorted(timers.items()):
        prog = f"{td.done_checks}/{len(CHECKPOINTS)}"
        nxt, ov = td.status()
        err = '❌' if td.error else ''
        iid = tree.insert('', 'end', values=(st, td.serial, td.tech, prog, nxt, err, ov))
        if sel_st == st:
            tree.selection_set(iid)
    root.after(1000, refresh_tree)
refresh_tree()

# ── LOG AREA ───────────────────────────────────────────
logbox = tk.Listbox(root, height=5, font=('Consolas',10), bg='#f7f7f7')
logbox.pack(fill='x')
def log(msg):
    ts = datetime.now().strftime('%H:%M:%S')
    activity_log.insert(0, f"[{ts}] {msg}")
    logbox.delete(0, tk.END)
    for e in activity_log[:5]:
        logbox.insert(tk.END, e)

# ── EXCEL I/O ──────────────────────────────────────────
def write_excel(row):
    wb = load_workbook(LOG_PATH)
    ws = wb.active
    ws.append(row)
    wb.save(LOG_PATH)

# ── SCAN DIALOG ───────────────────────────────────────
def scan(prompt):
    win = tk.Toplevel(root)
    win.title(prompt)
    win.transient(root); win.grab_set(); win.focus_force()
    ttk.Label(win, text=prompt).pack(pady=10)
    var = tk.StringVar()
    ent = ttk.Entry(win, textvariable=var)
    ent.pack(pady=10); ent.focus()
    def ok(e=None):
        if not var.get().strip():
            messagebox.showerror('Invalid', prompt)
            return
        win.grab_release(); win.destroy()
    ent.bind('<Return>', ok)
    ttk.Button(win, text='Confirm', command=ok).pack(pady=5)
    center(win, 300, 150)
    win.wait_window()
    return var.get().strip()

# ── DATA ENTRY POPUP ───────────────────────────────────
def prompt_user(serial, station, idx, sched_time, tech):
    def popup():
        prompt_lock.acquire(); beep()
        # 1) Serial verification
        user_serial = scan(f"Rescan Serial for Station {station}")
        if user_serial != serial:
            messagebox.showerror('Mismatch','Serial numbers don\'t match.')
            prompt_lock.release(); return
        # 2) Entry window
        win = tk.Toplevel(root)
        win.title(f'Station {station} – Check {idx}')
        win.transient(root); win.grab_set(); win.focus_force()
        win.geometry('600x400'); center(win,600,400)
        fields = [
            ('Voltage','digit'),('Current','digit'),
            ('Clamp Current','digit'),('Temperature','digit'),
            ('Control Light','toggle'),('Load Bank Light','toggle')
        ]
        entries = {}
        widgets = {}
        def save():
            row = [serial, station, idx, sched_time.strftime('%H:%M:%S'),
                   datetime.now().strftime('%H:%M:%S'), abs(int((datetime.now()-sched_time).total_seconds()))]
            for label, ftype in fields:
                if ftype == 'digit':
                    e1, e2 = widgets[label]
                    row.append(f"{e1.get().zfill(2)}.{e2.get().zfill(2)}")
                else:
                    row.append(widgets[label].get())
            row.append(tech)
            timers[station].error = any(widgets[l].get()=='OFF' for l in ('Control Light','Load Bank Light'))
            timers[station].done_checks += 1
            write_excel(row); log(f'✅ Logged Check {idx} for Station {station}')
            win.grab_release(); win.destroy(); prompt_lock.release()
        def ask(i=0):
            for w in win.winfo_children(): w.destroy()
            label, ftype = fields[i]
            ttk.Label(win, text=label).pack(pady=10)
            if ftype == 'digit':
                iv, dv = tk.StringVar(), tk.StringVar()
                vcmd = (win.register(lambda P: (P.isdigit() and len(P)<=2) or P==''), '%P')
                frm = ttk.Frame(win); frm.pack()
                e1 = ttk.Entry(frm, textvariable=iv, width=3, validate='key', validatecommand=vcmd)
                e2 = ttk.Entry(frm, textvariable=dv, width=3, validate='key', validatecommand=vcmd)
                e1.pack(side='left'); ttk.Label(frm, text='.').pack(side='left'); e2.pack(side='left')
                e1.focus(); e1.bind('<KeyRelease>', lambda e: e2.focus() if len(iv.get())==2 else None)
                def enter(e=None): entries[label]=(iv,dv); next_field()
                e2.bind('<Return>', enter); ttk.Button(win, text='Enter', command=enter).pack(pady=10)
            else:
                var = tk.StringVar(value='ON'); entries[label]=var
                f=ttk.Frame(win); f.pack(pady=10)
                ttk.Button(f, text='ON', command=lambda v=var: v.set('ON')).pack(side='left', expand=True, fill='x')
                ttk.Button(f, text='OFF', command=lambda v=var: v.set('OFF')).pack(side='left', expand=True, fill='x')
                ttk.Button(win, text='Enter', command=next_field).pack(pady=10)
        def next_field():
            if len(entries)<len(fields): ask(len(entries))
            else: review()
        def review():
            for w in win.winfo_children(): w.destroy()
            win.protocol('WM_DELETE_WINDOW', save)
            ttk.Label(win, text='Review All Data').pack(pady=10)
            grid = ttk.Frame(win); grid.pack()
            for r,(label,ftype) in enumerate(fields):
                ttk.Label(grid, text=label).grid(row=r, column=0, padx=5, pady=5)
                if ftype=='digit':
                    iv,dv=entries[label]
                    e1=ttk.Entry(grid,width=3); e1.insert(0,iv.get()); e1.grid(row=r,column=1)
                    ttk.Label(grid,text='.').grid(row=r,column=2)
                    e2=ttk.Entry(grid,width=3); e2.insert(0,dv.get()); e2.grid(row=r,column=3)
                    widgets[label]=(e1,e2)
                else:
                    var=entries[label]
                    cb=ttk.Combobox(grid,values=['ON','OFF'],textvariable=var,state='readonly')
                    cb.grid(row=r,column=1,columnspan=3,sticky='ew')
                    widgets[label]=var
            ttk.Button(win, text='Confirm & Save', command=save).pack(pady=15)
            center(win,600,400)
        ask(0)
    threading.Thread(target=popup, daemon=True).start()

# ── SCHEDULER & QUEUE ─────────────────────────────────
def schedule(serial, station, tech):
    td=TimerData(serial,station,tech); timers[station]=td
    def worker(idx,delay,start):
        time.sleep(delay)
        if not td.cancelled: prompt_q.put((serial,station,idx,start+timedelta(seconds=sum(CHECKPOINTS[:idx])),tech))
    for i,sec in enumerate(CHECKPOINTS,start=1):
        threading.Thread(target=worker,args=(i,sec,td.start),daemon=True).start()
    log(f'⏱ Started Station {station} ({serial})')
threading.Thread(target=lambda:[prompt_user(*prompt_q.get()) for _ in iter(int,1)],daemon=True).start()

# ── BUTTON BAR ────────────────────────────────────────
bar=ttk.Frame(root); bar.pack(fill='x',pady=5)

ttk.Label(bar,text='Technician:').pack(side='left',padx=5)
ttk.Entry(bar,textvariable=tech_var).pack(side='left',padx=5)

def start():
    if not tech_var.get().strip(): messagebox.showerror('Tech','Enter technician name'); return
    serial=scan('Scan Serial Number')
    while True:
        st_str=scan(f'Scan Station Number (1–{NUM_STATIONS})')
        if not st_str.isdigit(): messagebox.showerror('Invalid','Station must be a number'); continue
        st=int(st_str)
        if st<1 or st>NUM_STATIONS: messagebox.showerror('Invalid',f'Station must be 1–{NUM_STATIONS}'); continue
        break
    if st in timers and timers[st].done_checks<len(CHECKPOINTS) and not timers[st].cancelled: messagebox.showerror('Busy',f'Station {st} already active'); return
    schedule(serial,st,tech_var.get().strip())

def cancel():
    sel=tree.selection();
    if not sel: messagebox.showerror('Select','Select a station'); return
    st=tree.item(sel[0])['values'][0]
    if not messagebox.askyesno('Cancel',f'Cancel Station {st}?'): return
    td=timers.pop(st); td.cancelled=True
    wb=load_workbook(LOG_PATH); ws=wb.active
    for r,row in enumerate(list(ws.iter_rows(min_row=2)),start=2):
        if row[1].value==st: ws.delete_rows(r)
    wb.save(LOG_PATH); log(f'❌ Cancelled Station {st}')

ttk.Button(bar,text='Start',command=start).pack(side='left',padx=5)
ttk.Button(bar,text='Cancel',command=cancel).pack(side='left',padx=5)

root.mainloop()
